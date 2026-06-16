# -*- coding: utf-8 -*-
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re
import random
import tempfile
from datetime import datetime
from pathlib import Path
import os

app = Flask(__name__)
CORS(app)

# ==================== 词库 ====================
HOT_MODIFIER_KEYWORDS = {
    "face_male": ["眼神深邃如古井", "目光如炬似能洞穿人心", "剑眉星目英气逼人", "面如冠玉肤若凝脂", "冷峻孤傲如冰山", "不怒自威自带王者气场"],
    "face_female": ["眸若秋水含情脉脉", "眉目如画精致绝伦", "肤若凝脂吹弹可破", "倾国倾城绝世容颜", "超凡脱俗不染凡尘", "温婉可人如江南水乡"],
    "hair_dynamic": ["发丝飘逸如风", "无风自动宛如仙人", "光泽柔顺如丝绸", "青丝如瀑倾泻而下", "墨发飞扬带着几分狂放"],
    "cloth_texture": ["衣袂翻飞如仙鹤展翅", "流光溢彩华贵非凡", "质感高级一看就价值不菲", "剪裁得体完美衬托身形", "面料考究触感柔软"],
    "light_aura": ["周身光晕如神祇降临", "气场环绕如实质", "逆光剪影宛如画卷", "霞光万道瑞气千条", "光影交错层次分明"],
    "momentum": ["气势如虹直冲云霄", "气场全开压制全场", "王者降临众生俯首", "睥睨众生目空一切", "霸气侧漏令人不敢直视"],
    "micro_expression": ["眼角微微泛红强忍泪意", "嘴角微微上扬带着一丝不易察觉的笑意", "眉头紧锁仿佛在思考什么难题", "睫毛轻轻颤动暴露了内心的紧张"],
    "micro_movement": ["指尖在桌面上轻轻敲击", "拇指和食指无意识地摩挲着", "肩膀微微下沉似乎放下了什么重担", "脊背挺直如松"],
}

QUALITY_KEYWORDS = {
    "base_cn": "8K超高清，电影级画质，光影细腻如油画，史诗感扑面而来，大师级作品构图",
    "base_en": "8K, cinematic quality, volumetric lighting, highly detailed, masterpiece",
    "composition_cn": "黄金分割构图，主体突出，背景虚化适度",
}

ENGLISH_PARAMS = {
    "mj_ratio": "--ar 9:16",
    "mj_version": "--v 6",
    "mj_style": "--style raw",
}

STYLE_PRESETS = {
    "现代都市": {
        "scene_main": "极简风格的高档办公室，落地窗外是城市天际线",
        "scene_detail": "阳光透过百叶窗斜斜照进来，光影分明，空气中微尘飞舞",
        "age_f": "28岁", "age_m": "28岁",
        "face_f": "精致妆容，干练气质", "face_m": "清爽短发，面容冷峻，剑眉星目",
        "hair_f": "长发披肩或干练马尾，发丝在光线下泛着微光", "hair_m": "短发清爽，额前有几缕碎发，发梢带光",
        "cloth_f": "时尚职业装，剪裁合体，面料考究", "cloth_m": "合身西装，面料考究，领带一丝不苟，衣袂翻飞",
        "tone": "通透现代，略带冷色调，干净明亮",
        "light": "冷白顶灯与窗外暖阳形成冷暖对比，面部轮廓分明",
        "atmosphere": "专业干练，都市精英感，空气中仿佛弥漫着无形的压力",
        "tag": "都市写实，电影质感",
        "negative_cn": "古装，仙侠，卡通，模糊",
    },
    "玄幻仙侠": {
        "scene_main": "云雾缭绕的仙山，悬浮山峰，灵气光点飘散",
        "scene_detail": "金光洒落，霞光万道，灵泉飞瀑，仙鹤飞过，空气中弥漫着灵气",
        "age_f": "18岁", "age_m": "20岁",
        "face_f": "仙气飘渺，超凡脱俗，眉目如画", "face_m": "气质出尘，剑眉星目，面如冠玉",
        "hair_f": "如瀑青丝垂至腰际，发间点缀玉簪，青丝如瀑", "hair_m": "墨发高束成发髻，一根玉簪固定，几缕碎发随风飘动",
        "cloth_f": "流光仙裙，衣袂层叠，裙摆绣着精致花纹，衣袂翻飞", "cloth_m": "白色道袍，金丝镶边，腰系玉带，袖口宽大",
        "tone": "梦幻流光，青金色调，仙气飘渺",
        "light": "灵气光晕环绕周身，局部逆光勾勒轮廓，仿佛沐浴在圣光之中",
        "atmosphere": "仙气飘渺，超凡脱俗，如坠仙境",
        "tag": "玄幻仙侠，古风CG",
        "negative_cn": "现代，卡通，文字，西方元素",
    },
    "古风穿越": {
        "scene_main": "亭台楼阁，红墙绿瓦，庭院深深，樱花飘落",
        "scene_detail": "古琴悠扬，珠帘轻晃，香炉紫烟袅袅，檀香弥漫",
        "age_f": "20岁", "age_m": "22岁",
        "face_f": "温婉可人，眉目如画，肤若凝脂", "face_m": "温润如玉，风度翩翩，面如冠玉",
        "hair_f": "发髻配珠钗，青丝如瀑，步摇轻晃", "hair_m": "长发束冠，玉簪固定，发带飘飞",
        "cloth_f": "精美汉服，刺绣长裙，丝绸质感，衣袂飘飘", "cloth_m": "锦袍长衫，玉带束腰，袖口绣着云纹",
        "tone": "暖黄水墨感，复古柔和，古韵悠然",
        "light": "柔和自然光，晨曦或黄昏，光影柔和如画",
        "atmosphere": "古韵悠然，诗意盎然，如入画卷",
        "tag": "写实古风，中国传统美学",
        "negative_cn": "现代元素，科幻，卡通，西方",
    },
    "悬疑惊悚": {
        "scene_main": "昏暗的废弃走廊，墙皮剥落，露出红砖",
        "scene_detail": "灯光忽明忽暗，发出滋滋电流声，墙角有暗红色痕迹，霉味弥漫",
        "age_f": "25岁", "age_m": "28岁",
        "face_f": "紧张惊恐，脸色苍白，眼神闪躲", "face_m": "阴郁深沉，眼神锐利，嘴角挂着一丝诡异的笑",
        "hair_f": "长发散乱，略显狼狈，几缕头发粘在脸上", "hair_m": "短发凌乱，像是不久前经历过一场搏斗",
        "cloth_f": "深色风衣，衣角沾着灰尘", "cloth_m": "深色夹克，领口竖起，神秘莫测",
        "tone": "冷灰低饱和，高对比度，暗部细节丰富",
        "light": "硬光强阴影，局部光源如手电筒光，造成强烈的明暗对比",
        "atmosphere": "紧张压抑，不寒而栗，仿佛有什么在暗处窥视",
        "tag": "悬疑惊悚，黑暗电影风格",
        "negative_cn": "明亮，鲜艳，喜剧，温馨",
    },
    "甜宠恋爱": {
        "scene_main": "暖色咖啡馆，或樱花树下，柔光滤镜",
        "scene_detail": "阳光透过落地窗洒在木质桌面上，桌上一杯冒着心形拉花的拿铁，窗外樱花飘落",
        "age_f": "20岁", "age_m": "22岁",
        "face_f": "甜美可爱，笑容治愈，眼角带着笑意", "face_m": "阳光帅气，笑容温暖，眼神干净",
        "hair_f": "长发披肩，发尾微卷，温柔甜美", "hair_m": "清爽短发，刘海微微翘起，干净阳光",
        "cloth_f": "浅色针织连衣裙，温柔质感，软糯舒适", "cloth_m": "干净的白衬衫，袖口微微挽起，阳光干净",
        "tone": "暖色柔和，粉色系，低对比度，奶油质感",
        "light": "柔光背光，梦幻光晕，像加了柔焦滤镜",
        "atmosphere": "温馨浪漫，甜蜜治愈，空气中仿佛有甜甜的味道",
        "tag": "恋爱甜宠，韩剧质感",
        "negative_cn": "阴暗，冷色调，暴力，恐怖",
    },
    "武侠江湖": {
        "scene_main": "云雾缭绕的山巅，古松盘虬，远处群山如黛",
        "scene_detail": "剑气纵横，衣袂飘飘，江湖气息浓郁，风声萧萧",
        "age_f": "22岁", "age_m": "24岁",
        "face_f": "英气逼人，侠女风范，眼神坚定", "face_m": "侠气凛然，眼神坚定，不怒自威",
        "hair_f": "高马尾或侠女发髻，干净利落", "hair_m": "束发或披肩发，侠客风，发带飘飞",
        "cloth_f": "劲装或侠女装，干练英气，腰带束身", "cloth_m": "武侠劲装，潇洒飘逸，袖口紧束",
        "tone": "水墨感，青绿色调，江湖气息",
        "light": "自然光+雾气散射，光影柔和如画",
        "atmosphere": "江湖气息，快意恩仇，侠客风范",
        "tag": "武侠江湖，古风动作",
        "negative_cn": "现代，宫廷，卡通，西方",
    },
}

EN_STYLE_DETAILS = {
    "现代都市": {
        "tag": "Modern urban cinematic, realistic",
        "scene_main": "Minimalist high-end office with floor-to-ceiling windows overlooking the city skyline",
        "scene_detail": "Sunlight streaming through venetian blinds",
        "face_m": "Sharp jawline, clean-cut short hair, cold and handsome",
        "face_f": "Refined makeup, professional and capable aura",
        "hair_m": "Short and tidy, a few strands casually falling over forehead",
        "hair_f": "Long hair tied in a neat ponytail, strands glowing",
        "cloth_m": "Well-fitted suit, premium fabric, tie perfectly knotted",
        "cloth_f": "Chic business attire, tailored silhouette, luxurious fabric",
        "tone": "Crisp and modern, slightly cool tone",
        "light": "Mixed cool and warm light creating contrast",
        "atmosphere": "Professional and efficient, elite urban vibe",
        "negative": "ancient, cartoon, blurry, low quality",
    },
    "玄幻仙侠": {
        "tag": "Xianxia fantasy, mythical Chinese aesthetic, cinematic CG",
        "scene_main": "Misty sacred mountain with floating peaks, glowing spiritual particles",
        "scene_detail": "Golden light cascading, rainbow clouds, waterfalls with spiritual energy",
        "face_m": "Ethereal, sharp sword-like eyebrows, star-like eyes, jade-like skin",
        "face_f": "Immortal grace, transcendent beauty, delicate eyebrows",
        "hair_m": "Long black hair tied in a high bun with a jade hairpin",
        "hair_f": "Raven-black hair cascading to waist, jade ornaments",
        "cloth_m": "White Taoist robe with golden embroidery, jade belt",
        "cloth_f": "Luminous fairy dress, layered skirts with intricate embroidery",
        "tone": "Dreamy iridescent, cyan and gold tones",
        "light": "Spiritual aura glowing around the body, rim lighting",
        "atmosphere": "Mystical and sublime, otherworldly",
        "negative": "modern, cartoon, text, western elements, low quality",
    },
    "古风穿越": {
        "tag": "Historical Chinese costume drama, ancient aesthetics",
        "scene_main": "Ancient pavilions and towers, red walls and green tiles",
        "scene_detail": "Guqin melody, beaded curtains gently swaying, incense smoke",
        "face_m": "Gentle and refined, jade-like appearance, warm and scholarly",
        "face_f": "Gentle and lovely, picturesque eyebrows, flawless porcelain",
        "hair_m": "Long hair tied with a crown, jade hairpin, flowing ribbons",
        "hair_f": "Traditional bun with pearl hairpins, raven-black hair",
        "cloth_m": "Embroidered robe, jade belt, cloud patterns on sleeves",
        "cloth_f": "Exquisite Hanfu, embroidered long dress, silk texture",
        "tone": "Warm ochre, ink wash feeling, retro and soft",
        "light": "Soft natural light, dawn or dusk, painterly shadows",
        "atmosphere": "Ancient charm, poetic, like stepping into a painting",
        "negative": "modern elements, sci-fi, cartoon, western, low quality",
    },
}

DEFAULT_EN_STYLE = {
    "tag": "Cinematic, high quality",
    "scene_main": "Beautiful cinematic scene",
    "scene_detail": "Detailed environment with atmospheric lighting",
    "face_m": "Handsome, expressive, determined",
    "face_f": "Beautiful, expressive, graceful",
    "hair_m": "Well-styled hair, slight movement",
    "hair_f": "Silky hair, flowing gently",
    "cloth_m": "Elegant costume, fine texture",
    "cloth_f": "Elegant dress, fine texture",
    "tone": "Cinematic color grading",
    "light": "Natural lighting with soft shadows",
    "atmosphere": "Immersive atmospheric mood",
    "negative": "low quality, blurry, distorted, text, watermark",
}

EN_HOT_MODIFIERS = {
    "face_male": ["piercing eyes that see through souls", "commanding presence with cold gaze", "handsome and heroic"],
    "face_female": ["eyes like autumn water, full of emotion", "flawless porcelain skin", "ethereal beauty"],
    "hair_dynamic": ["hair flowing gracefully", "silky strands shimmering", "wild locks dancing"],
    "cloth_texture": ["robe fluttering like a crane", "luxurious fabric with subtle sheen", "tailored perfectly"],
    "momentum": ["domineering aura suppressing the scene", "king returning, all bow", "oppressive energy radiating"],
    "micro_expression": ["slightly reddened eyes holding back tears", "subtle smile hinting hidden joy", "furrowed brows in deep thought"],
    "micro_movement": ["fingertips gently tapping", "thumb rubbing unconsciously", "shoulders slightly lowered"],
}

EMOTION_KEYWORDS = {
    "轻蔑": ["轻蔑", "蔑视", "不屑"], "冰冷": ["冰冷", "冷酷", "冷漠"], "嚣张": ["嚣张", "张狂", "狂妄"],
    "冷笑": ["冷笑", "讥笑", "嘲讽"], "愤怒": ["愤怒", "生气", "怒火", "暴怒"], "悲伤": ["悲伤", "难过", "伤心", "悲痛"],
    "喜悦": ["喜悦", "开心", "高兴", "欢喜"], "恐惧": ["恐惧", "害怕", "惊恐", "畏惧"], "震惊": ["震惊", "惊讶", "惊愕", "诧异"],
    "温柔": ["温柔", "柔和", "体贴", "温情"], "坚定": ["坚定", "决心", "意志"], "自然": ["平静", "淡然", "自然"],
}

EMOTION_MAP = {k: v[0] for k, v in EMOTION_KEYWORDS.items()}

MOVE_DESCRIPTIONS = {
    "固定": "镜头固定不动", "摇": "镜头上下左右摇动", "移": "镜头横向移动",
    "跟": "镜头跟随主体", "推": "镜头向前推进", "拉": "镜头向后拉远",
    "升降": "镜头上下升降", "环绕": "镜头360度旋转",
}

SHOT_CONFIG = {
    "对话日常": {"sizes": ["中景", "近景", "特写"], "size_w": [30,45,25],
                "moves": ["固定","缓推","固定"], "move_w": [50,30,20], "dur": (2.5,5.0)},
    "动作场面": {"sizes": ["远景","全景","中景","近景"], "size_w": [15,25,25,35],
                "moves": ["跟","手持","快切","摇"], "move_w": [25,25,25,25], "dur": (1.5,4.0)},
    "情感戏": {"sizes": ["中景","近景","特写","大特写"], "size_w": [20,35,30,15],
                "moves": ["固定","缓推","缓拉"], "move_w": [40,35,25], "dur": (4.0,8.0)},
    "高潮场面": {"sizes": ["全景","中景","近景","特写"], "size_w": [20,30,30,20],
                "moves": ["缓推","环绕","升降"], "move_w": [35,35,30], "dur": (5.0,10.0)},
}

SPECIAL_ROLES = {"系统音", "旁白", "画外音", "系统", "字幕", "提示音", "音效"}

# ==================== 工具函数 ====================
def get_random_hot_modifier(category):
    if category in HOT_MODIFIER_KEYWORDS and HOT_MODIFIER_KEYWORDS[category]:
        return random.choice(HOT_MODIFIER_KEYWORDS[category])
    return ""

def get_en_hot_modifier(category):
    if category in EN_HOT_MODIFIERS and EN_HOT_MODIFIERS[category]:
        return random.choice(EN_HOT_MODIFIERS[category])
    return ""

def weighted_choice(options, weights):
    if not options:
        return ""
    total = sum(weights)
    r = random.random() * total
    cum = 0
    for opt, w in zip(options, weights):
        cum += w
        if r <= cum:
            return opt
    return options[0]

def get_shot_config_for_type(scene_type, importance, content=""):
    cfg = SHOT_CONFIG.get(scene_type, SHOT_CONFIG["对话日常"])
    size = weighted_choice(cfg["sizes"], cfg["size_w"])
    if importance in ["关键", "高潮"]:
        if size in ["远景", "大远景"]:
            size = "全景"
    move = weighted_choice(cfg["moves"], cfg["move_w"])
    duration = random.uniform(*cfg["dur"])
    content_len = len(content) if content else 0
    if content_len > 50:
        duration += 1.5
    elif content_len > 30:
        duration += 1.0
    duration = max(1.5, min(duration, 12.0))
    return {"size": size, "move": move, "duration": round(duration, 1)}

def parse_script(text):
    rows = []
    lines = text.strip().split('\n')
    for line in lines:
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        parts = re.split(r'[,，]+', line)
        if len(parts) >= 3:
            scene = parts[0].strip()
            role = parts[1].strip()
            content = '，'.join(parts[2:])
        elif len(line) > 5:
            content = line
            role = "未知角色"
            scene = "未知场景"
        else:
            continue
        
        emotion_match = re.search(r'[【{](.+?)[】}]', content)
        if emotion_match:
            emotion_text = emotion_match.group(1)
            emotion = "自然"
            for e, kw in EMOTION_KEYWORDS.items():
                if emotion_text in kw:
                    emotion = e
                    break
            content = re.sub(r'[【{].+?[】}]', '', content).strip()
        else:
            emotion = "自然"
        
        if not scene:
            scene = "未知场景"
        if not role:
            role = "未知角色"
        if not content:
            content = line
        
        rows.append({
            "scene": scene[:80],
            "role": role[:50],
            "content": content[:500],
            "emotion": emotion,
        })
    return rows

def build_core_prompt_cn(row, shot, style, age, is_female):
    tag = style.get('tag', '电影质感')
    shot_size = shot['size']
    move_desc = MOVE_DESCRIPTIONS.get(shot.get('move', '固定'), '镜头固定')
    
    if row['role'] in SPECIAL_ROLES:
        body = f"{row['role']}（仅音效，无实体形象）"
    else:
        gender = "女性" if is_female else "男性"
        base_face = style.get("face_f" if is_female else "face_m", "")
        hot_face = get_random_hot_modifier("face_female" if is_female else "face_male")
        face_desc = f"{base_face}，{hot_face}" if hot_face else base_face
        
        hair = style.get("hair_f" if is_female else "hair_m", "")
        hot_hair = get_random_hot_modifier("hair_dynamic")
        hair_desc = f"{hair}，{hot_hair}" if hot_hair else hair
        
        cloth = style.get("cloth_f" if is_female else "cloth_m", "")
        hot_cloth = get_random_hot_modifier("cloth_texture")
        cloth_desc = f"{cloth}，{hot_cloth}" if hot_cloth else cloth
        
        micro_expr = get_random_hot_modifier("micro_expression")
        micro_move = get_random_hot_modifier("micro_movement")
        hot_momentum = get_random_hot_modifier("momentum")
        
        body = f"{age}{gender}，{face_desc}。{hair_desc}。身着{cloth_desc}"
        if micro_expr:
            body += f"。{micro_expr}"
        if micro_move:
            body += f"。{micro_move}"
        if hot_momentum:
            body += f"。{hot_momentum}"
    
    scene = style.get('scene_main', '')
    scene_detail = style.get('scene_detail', '')
    emotion_cn = EMOTION_MAP.get(row['emotion'], row['emotion'])
    negative = f"低质量，模糊，变形，畸形，多手指，坏手，{style.get('negative_cn', '')}，文字，水印，字幕"
    
    prompt = (f"【画面风格】{tag}，{QUALITY_KEYWORDS['base_cn']}。"
              f"【镜头】{shot_size}，{move_desc}。{QUALITY_KEYWORDS['composition_cn']}。"
              f"【主体】{body}。"
              f"【场景】{scene}。{scene_detail}。"
              f"【动作】{row['content']}，情绪{emotion_cn}。"
              f"【光线】{style.get('light', '')}。"
              f"【色调】{style.get('tone', '')}。"
              f"【氛围】{style.get('atmosphere', '')}。"
              f"【负面】{negative}")
    return prompt[:5000]

def build_core_prompt_en(row, shot, style_name, age, is_female):
    duration = shot.get('duration', 4.0)
    age_clean = age.replace("岁", "") if age and "岁" in age else (age if age else "30")
    if not age_clean.isdigit():
        age_clean = "30"
    role_name = row['role'].strip() if row['role'] and row['role'].strip() else "the character"
    
    en_style = EN_STYLE_DETAILS.get(style_name, DEFAULT_EN_STYLE.copy())
    
    emotion_map = {"愤怒": "angry", "喜悦": "joyful", "悲伤": "sad", "恐惧": "fearful",
                   "震惊": "shocked", "温柔": "gentle", "坚定": "determined", "自然": "natural"}
    emotion_en = emotion_map.get(row['emotion'], "natural")
    
    gender_word = "woman" if is_female else "man"
    face_base = en_style.get("face_f" if is_female else "face_m", "expressive face")
    hair_base = en_style.get("hair_f" if is_female else "hair_m", "well-styled hair")
    cloth_base = en_style.get("cloth_f" if is_female else "cloth_m", "elegant clothing")
    
    hot_face = get_en_hot_modifier("face_female" if is_female else "face_male")
    hot_hair = get_en_hot_modifier("hair_dynamic")
    hot_cloth = get_en_hot_modifier("cloth_texture")
    hot_momentum = get_en_hot_modifier("momentum")
    micro_expr = get_en_hot_modifier("micro_expression")
    micro_move = get_en_hot_modifier("micro_movement")
    
    subject_parts = [f"A {age_clean} year old {gender_word} named {role_name}"]
    subject_parts.append(face_base)
    if hot_face:
        subject_parts.append(hot_face)
    subject_parts.append(hair_base)
    if hot_hair:
        subject_parts.append(hot_hair)
    subject_parts.append(f"Wearing {cloth_base}")
    if hot_cloth:
        subject_parts.append(hot_cloth)
    if micro_expr:
        subject_parts.append(micro_expr)
    if micro_move:
        subject_parts.append(micro_move)
    if hot_momentum:
        subject_parts.append(hot_momentum)
    subject_text = ". ".join(subject_parts) + "."
    
    prompt = f"""{en_style.get('tag', 'Cinematic style')}, {QUALITY_KEYWORDS['base_en']}

Subject: {subject_text}

Action: performing the scene action naturally. Emotion: {emotion_en}.

Scene: {en_style.get('scene_main', 'Beautiful scene')}. {en_style.get('scene_detail', '')}
Lighting: {en_style.get('light', 'Natural lighting')}.
Tone: {en_style.get('tone', 'Cinematic color grading')}.
Atmosphere: {en_style.get('atmosphere', 'Immersive mood')}.

Shot: Medium shot, static camera, {duration:.1f} seconds.

Quality: 8K, cinematic, photorealistic, highly detailed textures, masterpiece.

Negative: {en_style.get('negative', 'text, watermark, low quality, blurry')}.

{ENGLISH_PARAMS['mj_ratio']} {ENGLISH_PARAMS['mj_version']} {ENGLISH_PARAMS['mj_style']}"""
    prompt = re.sub(r'[\u4e00-\u9fff]', '', prompt)
    return prompt[:4000]

def generate_storyboard(rows, style_name, scene_type):
    shots = []
    style = STYLE_PRESETS.get(style_name, STYLE_PRESETS["玄幻仙侠"])
    
    for idx, row in enumerate(rows):
        shot = get_shot_config_for_type(scene_type, "普通", row['content'])
        
        is_female = False
        female_suffixes = ["女", "妹", "姐", "娘", "夫人", "小姐", "仙子"]
        is_female = any(row['role'].endswith(suffix) for suffix in female_suffixes)
        
        age = style.get("age_f" if is_female else "age_m", "25岁")
        
        prompt_core_cn = build_core_prompt_cn(row, shot, style, age, is_female)
        prompt_core_en = build_core_prompt_en(row, shot, style_name, age, is_female)
        prompt_video_cn = prompt_core_cn + f"\n【时长推荐】{shot['duration']:.1f}秒"
        prompt_video_en = prompt_core_en + f"\nDuration recommendation: {shot['duration']:.1f} seconds"
        
        shots.append({
            "shot_no": idx + 1,
            "scene": row['scene'],
            "role": row['role'],
            "content": row['content'],
            "emotion": row['emotion'],
            "shot_size": shot['size'],
            "move": shot['move'],
            "duration": shot['duration'],
            "prompt_core_cn": prompt_core_cn,
            "prompt_core_en": prompt_core_en,
            "prompt_video_cn": prompt_video_cn,
            "prompt_video_en": prompt_video_en,
        })
    
    return shots

def export_excel(shots, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分镜表"
    
    headers = ["镜号", "场景", "角色", "台词", "情绪", "景别", "运镜", "时长(秒)", 
               "图片提示词(中文)", "图片提示词(英文)", "视频提示词(中文)", "视频提示词(英文)"]
    ws.append(headers)
    
    hf = Font(bold=True, color="FFFFFF", size=11)
    hf_fill = PatternFill(start_color="4472C4", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.font = hf
        cell.fill = hf_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    for s in shots:
        ws.append([s["shot_no"], s["scene"], s["role"], s["content"], s["emotion"],
                   s["shot_size"], s["move"], s["duration"],
                   s["prompt_core_cn"], s["prompt_core_en"],
                   s["prompt_video_cn"], s["prompt_video_en"]])
    
    widths = [5, 18, 12, 55, 12, 10, 12, 8, 120, 100, 120, 100]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border
    
    ws.freeze_panes = 'A2'
    wb.save(str(path))

# ==================== API 路由 ====================
@app.route('/health', methods=['GET'])
def health_check():
    return {"status": "ok", "message": "短剧分镜生成器 API 运行正常"}

@app.route('/generate', methods=['POST'])
def generate_api():
    try:
        data = request.get_json()
        if not data:
            return {"error": "请求体不能为空"}, 400
        
        script_text = data.get('script', '')
        if not script_text:
            return {"error": "剧本内容不能为空"}, 400
        
        style_name = data.get('style', '玄幻仙侠')
        scene_type = data.get('scene_type', '对话日常')
        
        if style_name not in STYLE_PRESETS:
            return {"error": f"不支持的风格: {style_name}"}, 400
        
        rows = parse_script(script_text)
        if not rows:
            return {"error": "无法识别剧本格式"}, 400
        
        shots = generate_storyboard(rows, style_name, scene_type)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_dir = tempfile.mkdtemp()
        excel_path = Path(temp_dir) / f"分镜表_{timestamp}.xlsx"
        export_excel(shots, excel_path)
        
        return send_file(
            excel_path,
            as_attachment=True,
            download_name=f"分镜表_{timestamp}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": f"生成失败: {str(e)}"}, 500

@app.route('/styles', methods=['GET'])
def get_styles():
    return {"styles": list(STYLE_PRESETS.keys())}

@app.route('/scene_types', methods=['GET'])
def get_scene_types():
    return {"scene_types": list(SHOT_CONFIG.keys())}

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)