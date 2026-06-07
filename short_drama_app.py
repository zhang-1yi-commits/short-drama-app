import streamlit as st
import random
import re
import json
import threading
import os
import tempfile
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Callable, Any
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection, Border, Side
from openpyxl.utils import get_column_letter
import io
import pandas as pd

# ==========================================
#  页面配置
# ==========================================
st.set_page_config(
    page_title="短剧分镜生成器 v6.0",
    page_icon="🎬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 自定义 CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #4472C4;
        text-align: center;
        margin-bottom: 1rem;
    }
    .stButton>button {
        border-radius: 8px;
        font-weight: bold;
    }
    .stProgress > div > div > div > div {
        background-color: #4472C4;
    }
</style>
""", unsafe_allow_html=True)

# ==========================================
#  配置常量
# ==========================================
CONFIG = {
    "app": {
        "name": "短剧分镜生成器",
        "version": "6.0",
        "wechat": "zzm_0623",
        "max_history": 50,
        "max_prompt_length": 800,
        "max_content_length": 200,
        "max_description_length": 300
    },
    "paths": {
        "role_memory": "short_drama_role_memory.json",
        "history": "short_drama_history.json",
        "templates": "prompt_templates.json"
    }
}

# 场景类型
SCENE_TYPES = ["对话室内", "快节奏搞笑", "动作悬疑", "古风穿越", "甜宠虐恋", "都市逆袭"]
DEFAULT_TYPE = "古风穿越"

# 风格与类型的推荐映射
STYLE_TYPE_MAP = {
    "玄幻修仙": "古风穿越",
    "现代都市": "都市逆袭",
    "古风穿越": "古风穿越",
    "悬疑惊悚": "动作悬疑",
    "科幻未来": "动作悬疑",
    "民国风情": "对话室内",
    "甜宠恋爱": "甜宠虐恋",
    "乡村乡土": "对话室内",
    "末日废土": "动作悬疑",
    "宫廷权谋": "古风穿越",
    "青春校园": "对话室内",
    "武侠江湖": "动作悬疑"
}

# 运镜描述
MOVE_DESCRIPTIONS = {
    "固定": "镜头固定，稳定观察",
    "缓推": "缓缓推进，拉近情感距离",
    "快切": "快速切换，制造紧张感",
    "跟": "跟随主体移动",
    "手持": "轻微晃动，真实感",
    "环绕": "环绕拍摄，全方位展现",
    "横移": "横向移动，展现空间",
    "甩镜": "快速甩动，冲击感",
    "升降": "升降镜头，展现层次",
    "推": "向前推进，强调情绪",
    "拉": "向后拉开，展现环境",
    "摇": "镜头摇摆，营造不安",
    "变焦": "急速变焦，强调冲击"
}

# 风格预设（12种）
STYLE_PRESETS = {
    "玄幻修仙": {
        "scene_main": "云雾缭绕的仙山广场，悬浮山峰",
        "scene_detail": "灵气光点漂浮，仙鹤飞过",
        "age_f": "18岁", "age_m": "20岁",
        "look_f": "五官精致，肤若凝脂", "look_m": "剑眉星目，面如冠玉",
        "cloth_f": "流光仙裙，衣带飘飘", "cloth_m": "白色道袍，金丝镶边",
        "tone": "梦幻流光，青金色调", "light": "灵气光晕，局部逆光",
        "atmosphere": "仙气飘渺", "tag": "玄幻修仙，古风CG",
        "negative": "现代，卡通，文字，水印"
    },
    "现代都市": {
        "scene_main": "极简办公室，落地窗",
        "scene_detail": "城市天际线，阳光透入",
        "age_f": "25岁", "age_m": "25岁",
        "look_f": "精致妆容，干练", "look_m": "清爽短发，自信",
        "cloth_f": "时尚职业装", "cloth_m": "合身西装",
        "tone": "通透冷色调", "light": "冷白顶灯+自然光",
        "atmosphere": "都市精英感", "tag": "都市写实，电影质感",
        "negative": "古装，卡通，模糊"
    },
    "古风穿越": {
        "scene_main": "亭台楼阁，红墙绿瓦",
        "scene_detail": "庭院深深，樱花飘落",
        "age_f": "20岁", "age_m": "22岁",
        "look_f": "温婉可人", "look_m": "温润如玉",
        "cloth_f": "汉服刺绣，发髻珠钗", "cloth_m": "锦袍长衫，玉带",
        "tone": "暖黄水墨感", "light": "柔和自然光，逆光",
        "atmosphere": "古韵悠然", "tag": "写实古风",
        "negative": "现代，科幻，霓虹"
    },
    "悬疑惊悚": {
        "scene_main": "昏暗走廊，阴影浓重",
        "scene_detail": "灯光忽明忽暗",
        "age_f": "25岁", "age_m": "28岁",
        "look_f": "紧张惊恐", "look_m": "阴郁深沉",
        "cloth_f": "深色风衣", "cloth_m": "深色夹克",
        "tone": "冷灰低饱和", "light": "硬光强阴影",
        "atmosphere": "紧张压抑", "tag": "悬疑惊悚",
        "negative": "明亮，鲜艳，喜剧"
    },
    "科幻未来": {
        "scene_main": "全息控制台，金属墙",
        "scene_detail": "霓虹灯光，未来城市",
        "age_f": "22岁", "age_m": "25岁",
        "look_f": "冷艳干练", "look_m": "冷静沉着",
        "cloth_f": "机能紧身衣", "cloth_m": "战术服",
        "tone": "冷蓝霓虹", "light": "体积光，发光边缘",
        "atmosphere": "科技感", "tag": "科幻写实",
        "negative": "古装，田园"
    },
    "民国风情": {
        "scene_main": "老上海街景",
        "scene_detail": "有轨电车，复古招牌",
        "age_f": "23岁", "age_m": "25岁",
        "look_f": "优雅端庄", "look_m": "绅士儒雅",
        "cloth_f": "旗袍，手推波纹", "cloth_m": "长风衣，礼帽",
        "tone": "复古暖棕", "light": "台灯暖光",
        "atmosphere": "复古怀旧", "tag": "民国写实",
        "negative": "现代建筑"
    },
    "甜宠恋爱": {
        "scene_main": "暖色咖啡馆",
        "scene_detail": "樱花树，柔光滤镜",
        "age_f": "20岁", "age_m": "22岁",
        "look_f": "甜美可爱", "look_m": "阳光帅气",
        "cloth_f": "浅色连衣裙", "cloth_m": "干净衬衫",
        "tone": "暖色柔和", "light": "柔光背光",
        "atmosphere": "温馨浪漫", "tag": "恋爱甜宠",
        "negative": "阴暗，暴力"
    },
    "乡村乡土": {
        "scene_main": "农家土屋，稻田",
        "scene_detail": "青山绿水，炊烟",
        "age_f": "30岁", "age_m": "32岁",
        "look_f": "朴实自然", "look_m": "憨厚老实",
        "cloth_f": "碎花布衣", "cloth_m": "旧式夹克",
        "tone": "暖色自然", "light": "自然日光",
        "atmosphere": "田园牧歌", "tag": "乡村写实",
        "negative": "都市，科幻"
    },
    "末日废土": {
        "scene_main": "废墟城市，断壁残垣",
        "scene_detail": "沙尘弥漫，锈迹斑斑",
        "age_f": "24岁", "age_m": "26岁",
        "look_f": "坚毅疲惫", "look_m": "粗犷沧桑",
        "cloth_f": "破旧皮衣，护目镜", "cloth_m": "战术背心，防尘面罩",
        "tone": "暗黄灰蒙", "light": "侧光，阴影强烈",
        "atmosphere": "荒凉绝望", "tag": "末日废土",
        "negative": "明亮，繁华，干净"
    },
    "宫廷权谋": {
        "scene_main": "金碧辉煌大殿，龙椅",
        "scene_detail": "烛火摇曳，屏风纱幔",
        "age_f": "22岁", "age_m": "25岁",
        "look_f": "端庄华贵", "look_m": "威严深沉",
        "cloth_f": "凤冠霞帔，织金绣凤", "cloth_m": "龙袍玉带，十二章纹",
        "tone": "暗金红调", "light": "烛光映照，明暗对比",
        "atmosphere": "权谋压抑", "tag": "宫廷权谋",
        "negative": "现代，简约，明亮"
    },
    "青春校园": {
        "scene_main": "明亮教室，黑板",
        "scene_detail": "阳光透过窗户，课桌椅",
        "age_f": "17岁", "age_m": "17岁",
        "look_f": "清纯马尾", "look_m": "阳光校服",
        "cloth_f": "蓝白校服，百褶裙", "cloth_m": "蓝白校服，白衬衫",
        "tone": "清新明亮", "light": "自然光，侧逆光",
        "atmosphere": "青春洋溢", "tag": "青春校园",
        "negative": "暗黑，血腥，成人"
    },
    "武侠江湖": {
        "scene_main": "竹林深处，古亭",
        "scene_detail": "竹叶飘落，雾气缭绕",
        "age_f": "21岁", "age_m": "23岁",
        "look_f": "英气飒爽", "look_m": "侠气凛然",
        "cloth_f": "劲装短打，束腰", "cloth_m": "长衫佩剑，斗笠",
        "tone": "墨绿青调", "light": "林间光斑，逆光",
        "atmosphere": "江湖快意", "tag": "武侠江湖",
        "negative": "现代，科幻，枪械"
    }
}

# 镜头配置
SHOT_CONFIG = {
    "对话室内": {"sizes": ["中景", "近景", "特写"], "size_w": [30, 45, 25],
                 "angles": ["平视", "侧视"], "angle_w": [60, 40],
                 "moves": ["固定", "缓推", "固定"], "move_w": [40, 40, 20],
                 "dur": (2.5, 5.0)},
    "快节奏搞笑": {"sizes": ["特写", "近景", "中景"], "size_w": [45, 40, 15],
                   "angles": ["平视", "俯视"], "angle_w": [50, 50],
                   "moves": ["快切", "甩镜", "快切"], "move_w": [50, 30, 20],
                   "dur": (1.5, 3.5)},
    "动作悬疑": {"sizes": ["全景", "中景", "近景"], "size_w": [30, 45, 25],
                 "angles": ["平视", "动态跟踪"], "angle_w": [50, 50],
                 "moves": ["跟", "手持", "环绕"], "move_w": [40, 35, 25],
                 "dur": (2.0, 4.5)},
    "古风穿越": {"sizes": ["大全景", "全景", "中景"], "size_w": [30, 40, 30],
                  "angles": ["平视", "俯视"], "angle_w": [55, 45],
                  "moves": ["横移", "升降", "缓推"], "move_w": [35, 35, 30],
                  "dur": (3.5, 7.0)},
    "甜宠虐恋": {"sizes": ["近景", "特写", "中景"], "size_w": [45, 35, 20],
                  "angles": ["平视", "侧视"], "angle_w": [65, 35],
                  "moves": ["缓推", "环绕", "固定"], "move_w": [40, 30, 30],
                  "dur": (2.5, 5.5)},
    "都市逆袭": {"sizes": ["中景", "近景", "特写"], "size_w": [30, 50, 20],
                  "angles": ["平视", "俯视"], "angle_w": [60, 40],
                  "moves": ["快切", "固定", "推"], "move_w": [45, 35, 20],
                  "dur": (2.0, 4.5)},
}

# 特殊角色
SPECIAL_ROLES = {"系统音", "旁白", "画外音", "系统", "旁白音", "OS", "字幕", "背景音", "音效", "AI", "机器人", "旁白（男）", "旁白（女）"}

# 情绪映射
EMOTION_MAP = {
    "轻蔑": "眼神不屑，嘴角微扬，居高临下",
    "冰冷": "面无表情，眼神凌厉，气势逼人",
    "嚣张": "趾高气昂，表情张狂",
    "冷笑": "嘴角上扬，眼神嘲讽",
    "愤怒": "咬牙切齿，眼神凶狠",
    "悲伤": "眼眶泛红，泪水打转",
    "喜悦": "笑容灿烂，眼睛弯弯",
    "恐惧": "瞳孔放大，身体颤抖",
    "震惊": "目瞪口呆，难以置信",
    "温柔": "眼神柔和，轻声细语",
    "威严": "目光如炬，不怒自威",
    "羞涩": "脸红低头，眼神躲闪",
    "期待": "眼神发亮，身体前倾",
    "自然": "平静自然，不卑不亢",
    "绝望": "眼神空洞，面如死灰",
    "狂喜": "仰天大笑，手舞足蹈",
    "痛苦": "面容扭曲，冷汗直流",
    "嫉妒": "眼神阴鸷，咬牙切齿",
    "得意": "眉飞色舞，昂首挺胸",
    "困惑": "眉头紧锁，歪头思索"
}

EMOTION_KEYWORDS = {
    "轻蔑": ["轻蔑", "蔑视", "不屑", "看不起", "鄙夷"],
    "冰冷": ["冰冷", "冷酷", "冷漠", "冷淡", "无情"],
    "嚣张": ["嚣张", "张狂", "狂妄", "跋扈", "猖狂"],
    "冷笑": ["冷笑", "讥笑", "嘲讽", "嗤笑", "讥诮"],
    "愤怒": ["愤怒", "生气", "怒火", "暴怒", "怒不可遏"],
    "悲伤": ["悲伤", "难过", "伤心", "悲痛", "哀伤"],
    "喜悦": ["喜悦", "开心", "高兴", "欢喜", "欣喜"],
    "恐惧": ["恐惧", "害怕", "惊恐", "畏惧", "战栗"],
    "震惊": ["震惊", "惊讶", "惊愕", "惊诧", "错愕"],
    "温柔": ["温柔", "柔和", "体贴", "温情", "慈爱"],
    "威严": ["威严", "庄重", "霸气", "凛然", "肃穆"],
    "羞涩": ["羞涩", "害羞", "腼腆", "羞赧", "害臊"],
    "期待": ["期待", "盼望", "渴望", "希冀", "憧憬"],
    "绝望": ["绝望", "死灰", "万念俱灰", "心灰意冷"],
    "狂喜": ["狂喜", "欣喜若狂", "喜极而泣", "乐不可支"],
    "痛苦": ["痛苦", "剧痛", "折磨", "痛不欲生", "煎熬"],
    "嫉妒": ["嫉妒", "妒忌", "眼红", "醋意", "嫉恨"],
    "得意": ["得意", "洋洋得意", "沾沾自喜", "踌躇满志"],
    "困惑": ["困惑", "疑惑", "不解", "茫然", "迷糊"]
}

# ==========================================
#  数据类
# ==========================================

@dataclass
class InputRow:
    scene: str
    role: str
    content: str
    emotion: str
    importance: str = "普通"

@dataclass
class ShotUnit:
    shot_no: int = 0
    scene: str = ""
    type_text: str = ""
    style: str = ""
    role: str = ""
    content: str = ""
    emotion: str = ""
    importance: str = ""
    shot_size: str = ""
    angle: str = ""
    move: str = ""
    duration: float = 0.0
    description: str = ""
    emotion_tip: str = ""
    note: str = ""
    prompt_img: str = ""
    prompt_negative: str = ""
    prompt_vid: str = ""

# ==========================================
#  工具函数
# ==========================================

def get_data_dir() -> Path:
    """获取数据目录，云端环境使用临时目录"""
    # 检测是否在 Streamlit Cloud 环境
    if os.environ.get("STREAMLIT_SHARING") or not os.path.exists(str(Path.home() / "Desktop")):
        return Path(tempfile.gettempdir())
    return Path.home() / "Desktop"

def sanitize(text: str, max_len: int = 500) -> str:
    if not text:
        return ""
    # 使用原始字符串避免转义警告
    text = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9，。！？；：""''《》、\s\-]', '', text)
    if len(text) > max_len:
        text = text[:max_len - 3] + "..."
    return text.strip()

def extract_emotion(text: str) -> str:
    if not text:
        return "自然"
    for emotion, keywords in EMOTION_KEYWORDS.items():
        for kw in keywords:
            if kw in text:
                return emotion
    return "自然"

def is_special_role(role: str) -> bool:
    if not role:
        return False
    return role in SPECIAL_ROLES or any(role.endswith(s) for s in SPECIAL_ROLES)

def is_female(role: str) -> bool:
    female_indicators = ["女", "妹", "姐", "娘", "妈", "夫人", "小姐", "姑娘", "苏瑶", "小美", "妃", "公主", "皇后", "太后", "婢", "妾", "侍女", "仙女", "魔女", "妖女"]
    return any(ind in role for ind in female_indicators)

def weighted_choice(options: List[str], weights: List[int]) -> str:
    if not options:
        return ""
    if len(options) != len(weights):
        return options[0]
    total = sum(weights)
    r = random.random() * total
    cum = 0
    for opt, w in zip(options, weights):
        cum += w
        if r <= cum:
            return opt
    return options[0]

# ==========================================
#  管理器（线程安全版）
# ==========================================

class RoleMemory:
    def __init__(self):
        self._file = get_data_dir() / CONFIG["paths"]["role_memory"]
        self._data: Dict[str, Dict] = {}
        self._lock = threading.Lock()
        self._load()
    
    def _load(self):
        if self._file.exists():
            try:
                with open(self._file, 'r', encoding='utf-8') as f:
                    self._data = json.load(f)
            except Exception:
                self._data = {}
    
    def _save(self):
        try:
            with open(self._file, 'w', encoding='utf-8') as f:
                json.dump(self._data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
    
    def get(self, role: str, style: dict) -> Tuple[str, str]:
        with self._lock:
            if is_special_role(role):
                return "", ""
            if role in self._data:
                cloth = self._data[role].get("cloth", "")
                age = self._data[role].get("age", "")
                if cloth:
                    return cloth, age
            girl = is_female(role)
            age = style["age_f"] if girl else style["age_m"]
            cloth = style["cloth_f"] if girl else style["cloth_m"]
            return cloth, age
    
    def set(self, role: str, cloth: str, age: str):
        with self._lock:
            if is_special_role(role) or not role:
                return
            if role not in self._data:
                self._data[role] = {}
            self._data[role]["cloth"] = cloth
            self._data[role]["age"] = age
            self._save()
    
    def clear(self):
        with self._lock:
            self._data = {}
            self._save()
    
    def get_all(self) -> Dict[str, Dict]:
        with self._lock:
            return self._data.copy()
    
    def get_summary(self) -> str:
        with self._lock:
            if not self._data:
                return "暂无"
            items = [f"{r}({d.get('age', '')})" for r, d in list(self._data.items())[:5]]
            result = "，".join(items)
            if len(self._data) > 5:
                result += f" 等{len(self._data)}个"
            return result

class HistoryManager:
    def __init__(self):
        self._file = get_data_dir() / CONFIG["paths"]["history"]
        self._data: List[Dict] = []
        self._lock = threading.Lock()
        self._load()
    
    def _load(self):
        if self._file.exists():
            try:
                with open(self._file, 'r', encoding='utf-8') as f:
                    self._data = json.load(f)
            except Exception:
                self._data = []
    
    def _save(self):
        try:
            with open(self._file, 'w', encoding='utf-8') as f:
                json.dump(self._data[:CONFIG["app"]["max_history"]], f, ensure_ascii=False, indent=2)
        except Exception:
            pass
    
    def add(self, entry: Dict):
        with self._lock:
            self._data.insert(0, {"time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), **entry})
            self._save()
    
    def get_recent(self, limit: int = 10) -> List[Dict]:
        with self._lock:
            return self._data[:limit]
    
    def is_empty(self) -> bool:
        with self._lock:
            return len(self._data) == 0

class PromptTemplateManager:
    """提示词模板管理器 - 支持自定义模板"""
    def __init__(self):
        self._file = get_data_dir() / CONFIG["paths"]["templates"]
        self._templates: Dict[str, str] = {}
        self._lock = threading.Lock()
        self._load()
        self._defaults = {
            "img_default": "{tag}，景别：{size}，{body}，场景：{scene_main}，{scene_detail}，动作：{content}，{emotion}，氛围：{atmosphere}，色调：{tone}，光线：{light}，8k，电影质感",
            "vid_default": "{tag}，景别：{size}，{body}，场景：{scene_main}，动作：{content}，{emotion}，运镜：{move_desc}，氛围：{atmosphere}，色调：{tone}，光线：{light}，{duration}秒，禁止文字",
            "img_minimal": "{tag}，{scene_main}，{content}，{emotion}，{tone}，{light}，8k",
            "vid_cinematic": "电影级{tag}，{size}镜头，{body}在{scene_main}，{content}，{emotion}，{move_desc}，{tone}，电影光效，{duration}秒，禁止文字"
        }
    
    def _load(self):
        if self._file.exists():
            try:
                with open(self._file, 'r', encoding='utf-8') as f:
                    self._templates = json.load(f)
            except Exception:
                self._templates = {}
    
    def _save(self):
        try:
            with open(self._file, 'w', encoding='utf-8') as f:
                json.dump(self._templates, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
    
    def get_template(self, name: str) -> str:
        with self._lock:
            return self._templates.get(name, self._defaults.get(name, self._defaults["img_default"]))
    
    def save_template(self, name: str, template: str):
        with self._lock:
            self._templates[name] = template
            self._save()
    
    def list_templates(self) -> List[str]:
        with self._lock:
            return list(self._defaults.keys()) + list(self._templates.keys())
    
    def get_defaults(self) -> Dict[str, str]:
        return self._defaults.copy()

# ==========================================
#  核心逻辑
# ==========================================

def parse_script(text: str) -> List[InputRow]:
    rows = []
    for line in text.strip().split('\n'):
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        
        parts = re.split(r'[,，;；:：\s]+', line)
        if len(parts) < 3:
            continue
        
        scene = parts[0].strip()
        role = parts[1].strip()
        rest = ' '.join(parts[2:]).strip()
        
        emotion = "自然"
        importance = "普通"
        
        match = re.search(r'[【{](.+?)[】}]', rest)
        if match:
            emotion = extract_emotion(match.group(1))
            rest = re.sub(r'[【{].+?[】}]', '', rest).strip()
        else:
            emotion = extract_emotion(rest)
        
        if "【关键】" in line or "{关键}" in line:
            importance = "关键"
        elif "【高潮】" in line or "{高潮}" in line:
            importance = "高潮"
        
        rows.append(InputRow(
            scene=sanitize(scene, 50),
            role=sanitize(role, 30),
            content=sanitize(rest, CONFIG["app"]["max_content_length"]),
            emotion=emotion,
            importance=importance
        ))
    
    return rows

def get_shot_config(type_text: str, importance: str) -> dict:
    cfg = SHOT_CONFIG.get(type_text, SHOT_CONFIG["古风穿越"])
    size = weighted_choice(cfg["sizes"], cfg["size_w"])
    if importance in ["关键", "高潮"] and size in ["全景", "大全景"]:
        size = "近景"
    move = weighted_choice(cfg["moves"], cfg["move_w"])
    angle = weighted_choice(cfg["angles"], cfg["angle_w"])
    return {
        "size": size, "move": move, "angle": angle,
        "duration": random.uniform(*cfg["dur"])
    }

def build_prompt_img(row: InputRow, shot: dict, style: dict, cloth: str, age: str, 
                     template_mgr: PromptTemplateManager, template_name: str = "img_default") -> Tuple[str, str]:
    if is_special_role(row.role):
        body = f"主体：{row.role}（音效）"
    else:
        look = style["look_f"] if is_female(row.role) else style["look_m"]
        body = f"主体：{age}{look}{row.role}，{cloth}"
    
    emotion_desc = EMOTION_MAP.get(row.emotion, row.emotion)
    
    template = template_mgr.get_template(template_name)
    prompt = template.format(
        tag=style['tag'],
        size=shot['size'],
        body=body,
        scene_main=style['scene_main'],
        scene_detail=style.get('scene_detail', ''),
        content=row.content,
        emotion=emotion_desc,
        atmosphere=style['atmosphere'],
        tone=style['tone'],
        light=style['light'],
        duration=f"{shot['duration']:.1f}",
        move_desc=MOVE_DESCRIPTIONS.get(shot['move'], '镜头固定')
    )
    
    negative = f"低质量，模糊，变形，{style['negative']}，文字，水印"
    return prompt[:CONFIG["app"]["max_prompt_length"]], negative

def build_prompt_vid(row: InputRow, shot: dict, style: dict, cloth: str, age: str,
                     template_mgr: PromptTemplateManager, template_name: str = "vid_default") -> str:
    if is_special_role(row.role):
        body = f"主体：{row.role}"
    else:
        look = style["look_f"] if is_female(row.role) else style["look_m"]
        body = f"主体：{age}{look}{row.role}，{cloth}"
    
    emotion_desc = EMOTION_MAP.get(row.emotion, row.emotion)
    move_desc = MOVE_DESCRIPTIONS.get(shot['move'], "镜头固定")
    
    template = template_mgr.get_template(template_name)
    prompt = template.format(
        tag=style['tag'],
        size=shot['size'],
        body=body,
        scene_main=style['scene_main'],
        scene_detail=style.get('scene_detail', ''),
        content=row.content,
        emotion=emotion_desc,
        atmosphere=style['atmosphere'],
        tone=style['tone'],
        light=style['light'],
        duration=f"{shot['duration']:.1f}",
        move_desc=move_desc
    )
    
    return prompt[:CONFIG["app"]["max_prompt_length"]]

def generate_storyboard(rows: List[InputRow], style_name: str, scene_type: str, 
                        memory: RoleMemory, template_mgr: PromptTemplateManager,
                        img_template: str = "img_default", vid_template: str = "vid_default",
                        callback: Optional[Callable] = None) -> List[ShotUnit]:
    shots = []
    style = STYLE_PRESETS.get(style_name, STYLE_PRESETS["玄幻修仙"])
    
    for idx, row in enumerate(rows):
        if callback:
            callback(idx + 1, len(rows))
        
        shot = get_shot_config(scene_type, row.importance)
        shot['duration'] = round(shot['duration'] + len(row.content) * 0.02, 1)
        shot['duration'] = max(1.5, min(shot['duration'], 8.0))
        
        cloth, age = memory.get(row.role, style)
        prompt_img, prompt_neg = build_prompt_img(row, shot, style, cloth, age, template_mgr, img_template)
        prompt_vid = build_prompt_vid(row, shot, style, cloth, age, template_mgr, vid_template)
        
        emotion_short = EMOTION_MAP.get(row.emotion, row.emotion)[:20]
        description = f"{row.role}在{row.scene}。{row.content}。情绪：{emotion_short}。{shot['size']}，{shot['move']}。"
        
        shots.append(ShotUnit(
            shot_no=idx + 1,
            scene=row.scene,
            type_text=scene_type,
            style=style_name,
            role=row.role,
            content=row.content,
            emotion=row.emotion,
            importance=row.importance,
            shot_size=shot['size'],
            angle=shot['angle'],
            move=shot['move'],
            duration=shot['duration'],
            description=description[:CONFIG["app"]["max_description_length"]],
            emotion_tip=f"表现{emotion_short}",
            note="保持角色一致",
            prompt_img=prompt_img,
            prompt_negative=prompt_neg,
            prompt_vid=prompt_vid
        ))
        
        memory.set(row.role, cloth, age)
    
    return shots

def export_excel(shots: List[ShotUnit]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "分镜表"
    
    headers = ["镜号", "场景", "类型", "风格", "角色", "台词", "情绪", "重要性",
               "景别", "角度", "运镜", "时长", "画面描述", "情绪指导", 
               "拍摄备注", "即梦AI提示词", "负面提示词", "可灵AI提示词"]
    ws.append(headers)
    
    hf = Font(bold=True, color="FFFFFF", size=11)
    hf_fill = PatternFill(start_color="4472C4", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.font = hf
        cell.fill = hf_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    for s in shots:
        ws.append([s.shot_no, s.scene, s.type_text, s.style, s.role, s.content, 
                   s.emotion, s.importance, s.shot_size, s.angle, s.move, s.duration,
                   s.description, s.emotion_tip, s.note, s.prompt_img, 
                   s.prompt_negative, s.prompt_vid])
    
    widths = [5, 12, 10, 10, 10, 35, 9, 8, 8, 8, 8, 6, 50, 25, 25, 60, 50, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 55
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border
            cell.protection = Protection(locked=False)
    
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.protection.sheet = False
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def get_example_script() -> str:
    return """青云宗广场，苏瑶，你灵根已碎，修为尽废，今日我与你解除婚约。【轻蔑】
青云宗广场，林尘，今日之辱，他日百倍奉还。【冰冷】
青云宗广场，系统音，叮——反派继承系统已绑定，解锁1%记忆。【电子音】
杂役院，萧寒，废物，苏瑶让我来告诉你——滚出青云宗，越远越好。【嚣张】
杂役院，林尘，废你九成修为，算是还礼。【冷笑】"""

def get_example_script2() -> str:
    return """总裁办公室，顾总，这个项目必须拿下，不惜一切代价。【威严】
总裁办公室，秘书小李，顾总，对方已经拒绝了我们的提议。【恐惧】
咖啡厅，苏婉，顾总，我可以帮你，但有一个条件。【温柔】
咖啡厅，顾总，什么条件？你说。【期待】
天台，苏婉，我要你娶我。【羞涩】
天台，顾总，好，我答应你。【狂喜】"""

# ==========================================
#  Streamlit UI
# ==========================================

def init_session_state():
    if 'memory' not in st.session_state:
        st.session_state.memory = RoleMemory()
    if 'history' not in st.session_state:
        st.session_state.history = HistoryManager()
    if 'template_mgr' not in st.session_state:
        st.session_state.template_mgr = PromptTemplateManager()
    if 'generated_shots' not in st.session_state:
        st.session_state.generated_shots = None
    if 'last_style' not in st.session_state:
        st.session_state.last_style = "玄幻修仙"
    if 'last_type' not in st.session_state:
        st.session_state.last_type = "古风穿越"
    if 'script_input' not in st.session_state:
        st.session_state.script_input = get_example_script()

def main():
    init_session_state()
    
    # 标题
    st.markdown('<div class="main-header">🎬 短剧分镜生成器 v6.0</div>', unsafe_allow_html=True)
    st.caption("💡 格式：场景，角色，台词【情绪】 | 支持一键生成即梦/可灵AI提示词")
    
    # 侧边栏
    with st.sidebar:
        st.image("https://img.icons8.com/color/96/clapperboard.png", width=80)
        st.title("控制面板")
        st.caption(f"微信：{CONFIG['app']['wechat']}")
        st.markdown("---")
        
        # 角色记忆面板
        st.subheader("🧠 角色记忆")
        memory_summary = st.session_state.memory.get_summary()
        if memory_summary != "暂无":
            st.success(memory_summary)
        else:
            st.info("暂无角色记忆")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🗑️ 清除记忆", use_container_width=True):
                st.session_state.memory.clear()
                st.success("已清除角色记忆")
                st.rerun()
        with col2:
            if st.button("🔄 刷新", use_container_width=True):
                st.rerun()
        
        with st.expander("查看完整角色记忆"):
            all_mem = st.session_state.memory.get_all()
            if all_mem:
                for role, data in all_mem.items():
                    st.write(f"**{role}**：{data.get('age', '')}岁，{data.get('cloth', '')}")
            else:
                st.write("暂无记忆")
        
        st.markdown("---")
        
        # 历史记录
        st.subheader("📜 历史记录")
        if not st.session_state.history.is_empty():
            for h in st.session_state.history.get_recent(5):
                with st.container():
                    st.caption(f"{h['time']}")
                    st.write(f"{h.get('shots', 0)}个镜头 · {h.get('style', '')}")
        else:
            st.caption("暂无生成记录")
        
        st.markdown("---")
        st.caption("© 2025 短剧分镜生成器")
    
    # 主界面 - 剧本输入
    st.header("📝 剧本输入")
    
    # 快捷操作栏
    col_btn1, col_btn2, col_btn3, col_btn4 = st.columns(4)
    with col_btn1:
        if st.button("📋 加载示例（玄幻）", use_container_width=True):
            st.session_state.script_input = get_example_script()
            st.rerun()
    with col_btn2:
        if st.button("📋 加载示例（都市）", use_container_width=True):
            st.session_state.script_input = get_example_script2()
            st.rerun()
    with col_btn3:
        uploaded = st.file_uploader("📂 导入剧本", type=["txt"], label_visibility="collapsed")
        if uploaded:
            st.session_state.script_input = uploaded.read().decode('utf-8')
            st.rerun()
    with col_btn4:
        if st.button("🗑️ 清空", use_container_width=True):
            st.session_state.script_input = ""
            st.rerun()
    
    # 剧本输入区
    script_text = st.text_area(
        "输入剧本（每行格式：场景，角色，台词【情绪】）",
        value=st.session_state.script_input,
        height=250,
        placeholder="青云宗广场，苏瑶，你灵根已碎，修为尽废。【轻蔑】\n青云宗广场，林尘，今日之辱，他日百倍奉还。【冰冷】",
        key="script_input"
    )
    
    # 参数设置
    st.header("⚙️ 生成设置")
    
    col_set1, col_set2, col_set3, col_set4 = st.columns(4)
    
    with col_set1:
        style = st.selectbox(
            "🎨 风格",
            options=list(STYLE_PRESETS.keys()),
            index=list(STYLE_PRESETS.keys()).index(st.session_state.last_style),
            key="style_select"
        )
    
    with col_set2:
        recommended_type = STYLE_TYPE_MAP.get(style, DEFAULT_TYPE)
        type_index = SCENE_TYPES.index(recommended_type) if recommended_type in SCENE_TYPES else 0
        scene_type = st.selectbox(
            "🎬 场景类型",
            options=SCENE_TYPES,
            index=type_index,
            key="type_select"
        )
    
    with col_set3:
        img_template = st.selectbox(
            "📷 生图模板",
            options=["img_default", "img_minimal"],
            format_func=lambda x: {"img_default": "标准模板", "img_minimal": "精简模板"}.get(x, x),
            key="img_template"
        )
    
    with col_set4:
        vid_template = st.selectbox(
            "🎥 生视频模板",
            options=["vid_default", "vid_cinematic"],
            format_func=lambda x: {"vid_default": "标准模板", "vid_cinematic": "电影级模板"}.get(x, x),
            key="vid_template"
        )
    
    # 自定义模板扩展
    with st.expander("✏️ 自定义提示词模板（高级）"):
        st.info("使用占位符：{tag} {size} {body} {scene_main} {scene_detail} {content} {emotion} {atmosphere} {tone} {light} {duration} {move_desc}")
        
        col_t1, col_t2 = st.columns(2)
        with col_t1:
            custom_img = st.text_area("自定义生图模板", 
                value=st.session_state.template_mgr.get_template("img_default"),
                height=100)
            if st.button("保存生图模板"):
                st.session_state.template_mgr.save_template("custom_img", custom_img)
                st.success("已保存自定义生图模板！")
        with col_t2:
            custom_vid = st.text_area("自定义生视频模板",
                value=st.session_state.template_mgr.get_template("vid_default"),
                height=100)
            if st.button("保存生视频模板"):
                st.session_state.template_mgr.save_template("custom_vid", custom_vid)
                st.success("已保存自定义生视频模板！")
    
    # 生成按钮
    st.markdown("---")
    gen_col1, gen_col2, gen_col3 = st.columns([2, 1, 1])
    
    with gen_col1:
        generate_btn = st.button("🚀 一键生成分镜表", type="primary", use_container_width=True)
    
    # 生成逻辑
    if generate_btn:
        if not script_text.strip():
            st.error("⚠️ 请输入剧本内容")
            return
        
        rows = parse_script(script_text)
        if not rows:
            st.error("⚠️ 无法识别剧本格式，请使用：场景，角色，台词")
            return
        
        st.session_state.last_style = style
        st.session_state.last_type = scene_type
        
        # 进度显示
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        def update_progress(cur, total):
            progress = cur / total
            progress_bar.progress(min(progress, 1.0))
            status_text.text(f"⏳ 生成中... {cur}/{total} 个镜头")
        
        with st.spinner("正在生成分镜表..."):
            shots = generate_storyboard(
                rows, style, scene_type, 
                st.session_state.memory, 
                st.session_state.template_mgr,
                img_template, vid_template,
                update_progress
            )
        
        progress_bar.empty()
        status_text.empty()
        
        st.session_state.generated_shots = shots
        st.session_state.history.add({"shots": len(shots), "style": style, "type": scene_type})
        
        st.success(f"✅ 成功生成 **{len(shots)}** 个镜头！")
    
    # 显示结果
    if st.session_state.generated_shots:
        shots = st.session_state.generated_shots
        
        st.markdown("---")
        st.header(f"📊 分镜表预览（共 {len(shots)} 个镜头）")
        
        # 统计指标
        col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
        with col_stat1:
            st.metric("总镜头数", len(shots))
        with col_stat2:
            total_duration = sum(s.duration for s in shots)
            st.metric("预估总时长", f"{total_duration:.1f}秒")
        with col_stat3:
            unique_roles = len(set(s.role for s in shots))
            st.metric("角色数量", unique_roles)
        with col_stat4:
            avg_duration = total_duration / len(shots) if shots else 0
            st.metric("平均时长", f"{avg_duration:.1f}秒")
        
        # 导出按钮
        col_exp1, col_exp2, col_exp3 = st.columns(3)
        
        with col_exp1:
            excel_buffer = export_excel(shots)
            st.download_button(
                label="📥 下载 Excel 分镜表",
                data=excel_buffer,
                file_name=f"分镜表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        with col_exp2:
            prompts_text = "\n".join([f"【镜头{s.shot_no}】{s.role}\n{s.prompt_img}\n" for s in shots])
            st.download_button(
                label="📥 下载生图提示词",
                data=prompts_text,
                file_name=f"生图提示词_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                mime="text/plain",
                use_container_width=True
            )
        
        with col_exp3:
            vid_prompts_text = "\n".join([f"【镜头{s.shot_no}】{s.role}\n{s.prompt_vid}\n" for s in shots])
            st.download_button(
                label="📥 下载生视频提示词",
                data=vid_prompts_text,
                file_name=f"生视频提示词_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                mime="text/plain",
                use_container_width=True
            )
        
        # 数据表格
        st.subheader("📋 分镜详情")
        
        df_data = []
        for s in shots:
            df_data.append({
                "镜号": s.shot_no,
                "场景": s.scene,
                "角色": s.role,
                "台词": s.content,
                "情绪": s.emotion,
                "重要性": s.importance,
                "景别": s.shot_size,
                "运镜": s.move,
                "时长(秒)": s.duration,
                "画面描述": s.description,
                "即梦AI提示词": s.prompt_img[:80] + "..." if len(s.prompt_img) > 80 else s.prompt_img
            })
        
        df = pd.DataFrame(df_data)
        st.dataframe(df, use_container_width=True, height=400)
        
        # 详细卡片展示
        st.subheader("🎴 镜头卡片")
        
        filter_role = st.selectbox("筛选角色", options=["全部"] + sorted(list(set([s.role for s in shots]))))
        
        filtered_shots = [s for s in shots if filter_role == "全部" or s.role == filter_role]
        
        cols = st.columns(3)
        for idx, shot in enumerate(filtered_shots):
            with cols[idx % 3]:
                with st.container(border=True):
                    st.markdown(f"**🎬 镜头 {shot.shot_no}** | {shot.role}")
                    st.caption(f"{shot.shot_size} · {shot.move} · {shot.duration}秒 · {shot.importance}")
                    st.write(f"📝 {shot.content}")
                    st.write(f"😊 {shot.emotion}")
                    
                    with st.expander("查看提示词"):
                        st.text_area("即梦AI", shot.prompt_img, height=80, key=f"img_{shot.shot_no}")
                        st.text_area("可灵AI", shot.prompt_vid, height=80, key=f"vid_{shot.shot_no}")
                        st.text_area("负面提示词", shot.prompt_negative, height=60, key=f"neg_{shot.shot_no}")

if __name__ == "__main__":
    main()
